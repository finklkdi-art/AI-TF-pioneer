"""
Excel 내보내기 — 살아있는 수식(Source 25(b))을 유지.
- 단가 * 수량 = 금액 셀 수식 작성
- SUM() 으로 (A)(B)(C) 합계, VAT, 총합
- 파일명 prefix: BLUE_NINE_ (Source 37)
"""
from __future__ import annotations
import io
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

from .schemas import EstimateDocument

THIN = Side(border_style="thin", color="888888")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
HEADER_FILL = PatternFill("solid", fgColor="0B2A4A")
SECTION_FILL = PatternFill("solid", fgColor="E6EEF7")
TOTAL_FILL = PatternFill("solid", fgColor="DDE6F1")


def _set_header(cell, txt: str):
    cell.value = txt
    cell.font = Font(name="맑은 고딕", bold=True, color="FFFFFF", size=11)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.fill = HEADER_FILL
    cell.border = BORDER


def build_xlsx(doc: EstimateDocument) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "BLUE NINE 견적서"

    # 헤더
    ws.merge_cells("A1:H1")
    title = ws["A1"]
    title.value = f"BLUE NINE 광고 견적서 ({doc.category_l1.upper()} / {doc.category_l2.upper()})"
    title.font = Font(name="맑은 고딕", bold=True, size=16, color="0B2A4A")
    title.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 26

    meta = [
        ("광고주 (CLIENT)", doc.client or ""),
        ("캠페인 (Job Name)", doc.campaign or ""),
        ("Job No.", doc.job_no or ""),
        ("버전", doc.version_label),
        ("작성일", doc.issue_date or ""),
        ("처리 모드", doc.mode.upper()),
    ]
    for i, (k, v) in enumerate(meta, start=3):
        ws.cell(row=i, column=1, value=k).font = Font(bold=True)
        ws.cell(row=i, column=2, value=v)

    header_row = 10
    headers = ["섹션", "항목", "협력사/매체사", "단가", "수량", "금액(수식)", "신호등", "비고"]
    for j, h in enumerate(headers, start=1):
        _set_header(ws.cell(row=header_row, column=j), h)
    ws.row_dimensions[header_row].height = 22

    r = header_row + 1
    section_order = ["정가항목", "외주비", "대행수수료", "매체청구액", "매체지급액", "매체수수료"]
    rows_by_section = {sec: [row for row in doc.rows if row.section == sec] for sec in section_order}

    section_total_rows: dict[str, int] = {}
    for sec in section_order:
        bucket = rows_by_section[sec]
        if not bucket:
            continue
        # 섹션 표시
        sc = ws.cell(row=r, column=1, value=sec)
        sc.font = Font(bold=True, color="0B2A4A")
        sc.fill = SECTION_FILL
        for j in range(2, 9):
            ws.cell(row=r, column=j).fill = SECTION_FILL
            ws.cell(row=r, column=j).border = BORDER
        sc.border = BORDER
        r += 1
        start_amount_row = r
        for row in bucket:
            ws.cell(row=r, column=1, value="").border = BORDER
            ws.cell(row=r, column=2, value=row.item_name).border = BORDER
            ws.cell(row=r, column=3, value=row.vendor or "").border = BORDER
            ws.cell(row=r, column=4, value=row.unit_price).number_format = '#,##0'
            ws.cell(row=r, column=4).border = BORDER
            ws.cell(row=r, column=5, value=row.quantity).border = BORDER
            # 살아있는 수식 (Source 25b)
            cell_amount = ws.cell(row=r, column=6)
            cell_amount.value = f"=D{r}*E{r}"
            cell_amount.number_format = '#,##0'
            cell_amount.border = BORDER
            ws.cell(row=r, column=7, value={"green":"🟢","yellow":"🟡","red":"🔴"}.get(row.light,"")).border = BORDER
            ws.cell(row=r, column=8, value=row.note or "").border = BORDER
            r += 1
        end_amount_row = r - 1
        # 섹션 소계 — 알파벳 기호(A/B/C) 부여
        SECTION_LETTER = {"정가항목": "A", "외주비": "B", "대행수수료": "C"}
        letter = SECTION_LETTER.get(sec, "")
        sub_label = f"{sec} 소계 ({letter})" if letter else f"{sec} 소계"
        sub = ws.cell(row=r, column=2, value=sub_label)
        sub.font = Font(bold=True)
        sub.alignment = Alignment(horizontal="right")
        sub.fill = TOTAL_FILL
        for j in [1,2,3,4,5,6,7,8]:
            ws.cell(row=r, column=j).fill = TOTAL_FILL
            ws.cell(row=r, column=j).border = BORDER
        cell_sub = ws.cell(row=r, column=6)
        cell_sub.value = f"=SUM(F{start_amount_row}:F{end_amount_row})"
        cell_sub.font = Font(bold=True)
        cell_sub.number_format = '#,##0'
        section_total_rows[sec] = r
        r += 2

    # 합계 ((A)+(B)+(C) or (매체청구액))
    if doc.category_l1 == "production":
        ws.cell(row=r, column=2, value="(A) 정가합계 + (B) 외주비 + (C) 대행수수료").font = Font(bold=True)
        parts = []
        for sec in ("정가항목", "외주비", "대행수수료"):
            if sec in section_total_rows:
                parts.append(f"F{section_total_rows[sec]}")
        formula = "=" + ("+".join(parts) if parts else "0")
    else:
        formula = f"=F{section_total_rows.get('매체청구액', r-2)}"
        ws.cell(row=r, column=2, value="매체청구액 합계").font = Font(bold=True)
    cell_total = ws.cell(row=r, column=6, value=formula)
    cell_total.font = Font(bold=True, size=12)
    cell_total.number_format = '#,##0'
    cell_total.fill = TOTAL_FILL
    for j in range(1,9):
        ws.cell(row=r, column=j).fill = TOTAL_FILL
        ws.cell(row=r, column=j).border = BORDER
    r += 1
    # VAT 10%
    ws.cell(row=r, column=2, value="VAT (10%)").font = Font(bold=True)
    vat_cell = ws.cell(row=r, column=6, value=f"=ROUND(F{r-1}*0.1,0)")
    vat_cell.number_format = '#,##0'
    for j in range(1,9):
        ws.cell(row=r, column=j).border = BORDER
    r += 1
    ws.cell(row=r, column=2, value="청구금액 (VAT 포함)").font = Font(bold=True, size=12)
    grand = ws.cell(row=r, column=6, value=f"=F{r-2}+F{r-1}")
    grand.number_format = '#,##0'
    grand.font = Font(bold=True, size=12)
    for j in range(1,9):
        ws.cell(row=r, column=j).fill = TOTAL_FILL
        ws.cell(row=r, column=j).border = BORDER

    # 컬럼 폭
    widths = [12, 28, 18, 14, 10, 16, 10, 28]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Print area (Source 25a)
    ws.print_options.horizontalCentered = True
    ws.page_setup.orientation = "portrait"
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_margins.left = 0.4
    ws.page_margins.right = 0.4
    ws.page_margins.top = 0.6
    ws.page_margins.bottom = 0.6
    ws.oddHeader.center.text = "BLUE NINE 광고 견적서"
    ws.oddFooter.center.text = "Page &P / &N"

    bio = io.BytesIO()
    wb.save(bio)
    return bio.getvalue()
